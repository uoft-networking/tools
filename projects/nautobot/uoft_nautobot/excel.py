from django.forms import Form, FileField
from django.core.validators import FileExtensionValidator
from nautobot.apps.ui import TemplateExtension
from django.views.generic import View
from django.http import HttpResponse, StreamingHttpResponse
from django.shortcuts import render
from django.contrib.contenttypes.models import ContentType
from tempfile import NamedTemporaryFile

from nautobot.dcim.models import Interface, Device
from nautobot.extras.models import Status, Tag, Role
from nautobot.ipam.models import VLAN, VLANGroup, IPAddress
from nautobot.dcim.models.device_components import (
    InterfaceModeChoices,
    InterfaceTypeChoices,
    InterfaceStatusChoices,
)
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from glom import glom, register, Coalesce

from django.db.models import Manager, QuerySet
import nautobot.core.models.managers


def _register_manager_classes():
    # nautobot's Interface model contains a few dynamically-generated Manager classes
    # though they seem to be subclasses of django.db.models.Manager, glom doesn't
    # recognize them as such. We need to register them manually.
    # since these classes are dynamically generated at runtime, we can't just import them,
    # we need to dynamically aquire references to them.
    _example_intf = Interface()
    _ManyRelatedManager = _example_intf.tagged_vlans.__class__
    _GenericRelatedObjectManager = _example_intf.ip_addresses.__class__
    register(_ManyRelatedManager, iterate=lambda m: m.all())
    register(_GenericRelatedObjectManager, iterate=lambda m: m.all())

    register(Manager, iterate=lambda m: m.all())
    register(QuerySet, iterate=lambda qs: qs.all())
    register(
        nautobot.core.models.managers.TagsManager,
        iterate=lambda m: m.all(),
    )


class ExcelContext:
    def __init__(self, pk) -> None:
        self.device_obj = Device.objects.get(pk=pk)
        self.intfs: list[Interface] = list(self.device_obj.interfaces.all())  # pyright: ignore[reportAttributeAccessIssue]

        self.interface_types = InterfaceTypeChoices.as_dict()
        self.interface_types_inverse = {v: k for k, v in self.interface_types.items()}
        self.statuses = InterfaceStatusChoices.as_dict().values()
        self.vlan_modes = InterfaceModeChoices.values()
        _roles_ct = ContentType.objects.get(app_label="dcim", model="interface")
        self.valid_roles = list(Role.objects.filter(content_types=_roles_ct))

        #
        if self.device_obj.vlan_group:
            self.vlan_group = self.device_obj.vlan_group
        else:
            assert self.device_obj.name
            vlan_group_name = self.device_obj.name.partition("-")[2][:2]
            self.vlan_group = VLANGroup.objects.get(name=vlan_group_name)
        self.valid_vlans = VLAN.objects.filter(vlan_group=self.vlan_group)


def export_to_excel(pk):
    _register_manager_classes()
    ctx = ExcelContext(pk)

    wb = Workbook()
    assert wb.active
    ws: Worksheet = wb.active 
    ws.title = "Interfaces"

    def _vlan_to_string(vlan):
        if not vlan:
            return None
        return f"{vlan.vlan_group.name}/{vlan.name}({vlan.vid})"

    spec = [
        {
            "Name": "name",
            "Status": "status.name",
            "Role": Coalesce("role.name", default=""),
            "Label": "label",
            "Type": ("type", lambda t: ctx.interface_types[t]),
            "Enabled": "enabled",
            "LAG": Coalesce("lag.name", default=None),
            "Management Only": "mgmt_only",
            "Description": "description",
            "VLAN Mode": "mode",
            "Untagged VLAN": ("untagged_vlan", lambda v: _vlan_to_string(v)),
            "Tagged VLANs": (
                "tagged_vlans",
                ([lambda v: _vlan_to_string(v)], lambda lst: "\n".join(lst)),
            ),
            "Tags": ("tags", (["name"], lambda lst: "\n".join(lst))),
            "IP Addresses": (
                "ip_addresses",
                ([("address", str)], lambda lst: "\n".join(lst)),
            ),
            "Custom Fields": (
                "cf",
                (
                    ["name", "value"],
                    lambda lst: "\n".join([f"{k}: {v}" for k, v in lst]),
                ),
            ),
        }
    ]

    output = glom(ctx.intfs, spec)

    # need to calculate table size in excel
    # excel clumn and row numbers start at 1
    headers = tuple(output[0].keys())
    columns_by_name = {v: i + 1 for i, v in enumerate(headers)}
    width = len(headers)
    length = len(output) + 1  # extra +1 for header row

    # write headers
    ws.append(headers)

    # write data
    for row in output:
        ws.append(tuple(row.values()))

    # format as table
    tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(width)}{length}")
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Stop excel from treating intf names like 1/1/2 as dates
    for cell in ws["A"]:
        cell.number_format = "Text"

    # data validation
    dvws = wb.create_sheet("_data_validation_lists")
    dvws["A1"] = "statuses"
    for i, s in enumerate(ctx.statuses):
        dvws[f"A{i+2}"] = s
    dvws["B1"] = "interface types"
    for i, t in enumerate(ctx.interface_types.values()):
        dvws[f"B{i+2}"] = t
    dvws["C1"] = "vlan modes"
    for i, m in enumerate(ctx.vlan_modes):
        dvws[f"C{i+2}"] = m
    dvws["D1"] = "valid vlans"
    for i, v in enumerate(ctx.valid_vlans):
        dvws[f"D{i+2}"] = _vlan_to_string(v)
    dvws["E1"] = "roles"
    for i, r in enumerate(ctx.valid_roles):
        dvws[f"E{i+2}"] = r.name

    # boolean columns
    bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"')
    bool_dv.hide_drop_down = False
    ws.add_data_validation(bool_dv)
    for col in ["Enabled", "Management Only"]:
        col_letter = get_column_letter(columns_by_name[col])
        bool_dv.add(f"{col_letter}2:{col_letter}{length}")

    # LAG column
    lag_dv = DataValidation(type="list", formula1=f"'Interfaces'!$A$2:$A${length}", allow_blank=True)
    lag_dv.hide_drop_down = False
    ws.add_data_validation(lag_dv)
    col_letter = get_column_letter(columns_by_name["LAG"])
    lag_dv.add(f"{col_letter}2:{col_letter}{length}")

    # status
    status_dv = DataValidation(type="list", formula1=f"'_data_validation_lists'!$A$2:$A${len(ctx.statuses)+1}")
    status_dv.hide_drop_down = False
    ws.add_data_validation(status_dv)
    col_letter = get_column_letter(columns_by_name["Status"])
    status_dv.add(f"{col_letter}2:{col_letter}{length}")

    # interface types
    types_dv = DataValidation(
        type="list",
        formula1=f"'_data_validation_lists'!$B$2:$B${len(ctx.interface_types)+1}",
    )
    types_dv.hide_drop_down = False
    ws.add_data_validation(types_dv)
    col_letter = get_column_letter(columns_by_name["Type"])
    types_dv.add(f"{col_letter}2:{col_letter}{length}")

    # vlan modes
    modes_dv = DataValidation(
        type="list",
        formula1=f"'_data_validation_lists'!$C$2:$C${len(ctx.vlan_modes)+1}",
        allow_blank=True,
    )
    modes_dv.hide_drop_down = False
    ws.add_data_validation(modes_dv)
    col_letter = get_column_letter(columns_by_name["VLAN Mode"])
    modes_dv.add(f"{col_letter}2:{col_letter}{length}")

    # valid vlans
    vlans_dv = DataValidation(
        type="list",
        formula1=f"'_data_validation_lists'!$D$2:$D${len(ctx.valid_vlans)+1}",
        allow_blank=True,
    )
    vlans_dv.hide_drop_down = False
    ws.add_data_validation(vlans_dv)
    col_letter = get_column_letter(columns_by_name["Untagged VLAN"])
    vlans_dv.add(f"{col_letter}2:{col_letter}{length}")

    # roles
    roles_dv = DataValidation(
        type="list",
        formula1=f"'_data_validation_lists'!$E$2:$E${len(ctx.valid_roles)+1}",
        allow_blank=True,
    )
    roles_dv.hide_drop_down = False
    ws.add_data_validation(roles_dv)
    col_letter = get_column_letter(columns_by_name["Role"])
    roles_dv.add(f"{col_letter}2:{col_letter}{length}")

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        content = tmp.read()

    filename = f"{ctx.device_obj.name}-interfaces.xlsx"
    return filename, content


def import_from_excel(pk, file):
    _register_manager_classes()
    # TODO: Set default status of new ip addresses to Active
    yield "<p>Opening uploaded xlsx file...</p>"
    ctx = ExcelContext(pk)
    wb = openpyxl.load_workbook(file)
    ws = wb["Interfaces"]

    yield "<p>Reading data from xlsx file...</p>"

    # get the data back out of the excel file
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))

    def _parse_vlan(vlan):
        if not vlan:
            return None
        group, name = vlan.split("/")
        name = name.partition("(")[0]
        return VLAN.objects.get(vlan_group__name=group, name=name)

    def _parse_vlan_list(vlans: str):
        if not vlans:
            return []
        return [_parse_vlan(v) for v in vlans.split("\n")]

    def _parse_ipaddress_list(ips: str):
        if not ips:
            return []
        return [
            IPAddress.objects.get_or_create(address=ip, defaults=dict(status=Status.objects.get(name="Active")))[0]
            for ip in ips.split("\n")
        ]

    def _parse_tags(tags):
        if not tags:
            return []
        return [Tag.objects.get(name=tag) for tag in tags.split("\n")]

    yield "<p>Structuring and validating data...</p>"

    data = glom(
        data,
        (
            [
                {
                    "Payload": {
                        "name": "Name",
                        "defaults": {
                            "status": (
                                "Status",
                                lambda s: Status.objects.get(name=s),
                            ),
                            "label": ("Label", lambda lbl: lbl if lbl else ""),
                            "type": ("Type", lambda t: ctx.interface_types_inverse[t]),
                            "role": ("Role", lambda r: Role.objects.get(name=r) if r else None),
                            "enabled": "Enabled",
                            "lag": (
                                "LAG",
                                lambda lag: Interface.objects.get(name=lag, device_id=pk) if lag else None,
                            ),
                            "mgmt_only": "Management Only",
                            "description": ("Description", lambda d: d if d else ""),
                            "mode": ("VLAN Mode", lambda m: m if m else ""),
                            "untagged_vlan": (
                                "Untagged VLAN",
                                lambda v: _parse_vlan(v),
                            ),
                        },
                    },
                    "tagged_vlans": (
                        "Tagged VLANs",
                        lambda v: _parse_vlan_list(v),
                    ),
                    "ip_addresses": (
                        "IP Addresses",
                        lambda i: _parse_ipaddress_list(i),
                    ),
                    "cf": "Custom Fields",
                    "tags": ("Tags", lambda t: _parse_tags(t)),
                }
            ]
        ),
    )

    yield "<p>Writing data to database...</p>"

    for intf_data in data:
        payload = intf_data.pop("Payload")
        payload["device"] = ctx.device_obj
        intf = Interface.objects.update_or_create(**payload)[0]
        intf.ip_addresses.set(intf_data["ip_addresses"])
        intf.tagged_vlans.set(intf_data["tagged_vlans"])
        intf.tags.set(intf_data["tags"])
        if intf_data["cf"]:
            for k, v in intf_data["cf"].items():
                intf.custom_field_data[k] = v
        intf.validated_save()
        yield "."

    yield f"<p>Done! Please <a href='/dcim/devices/{pk}/interfaces'>click here</a> to see the changes.</p>"


class FileForm(Form):
    file = FileField(validators=[FileExtensionValidator(allowed_extensions=["xlsx"])])


class DeviceInterfacesExcel(View):
    def get(self, request, pk, *args, **kwargs):
        filename, content = export_to_excel(pk)
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    def post(self, request, pk, *args, **kwargs):
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]

            def excel_processing():
                try:
                    for chunk in import_from_excel(pk, file):
                        yield chunk
                except Exception as e:
                    yield "<p>Something went wrong: </p>"
                    for line in str(e).split("\n"):
                        yield f"<p>{line}</p>"
                    yield "<p>Please check the server logs for additional information</p>"

            return StreamingHttpResponse(excel_processing())  # pyright: ignore[reportArgumentType]
        else:
            # We've been sent here by the widget on the device page
            # but the data we've been sent here with failed server-side validation.
            # We need to render the form again, but with the errors from the form included
            form_html = render(request, "device_interfaces_excel.html", {"form": form, "pk": pk})
            return form_html

class DeviceInterfacesExcelWidget(TemplateExtension):
    """
    TemplateExtension class for adding buttons allowing to import/export a device's interface list from/to Excel.
    """

    model = "dcim.device"

    def right_page(self):
        form = FileForm()
        object = self.context["object"]
        form_html = self.render(
            "device_interfaces_excel.html",
            extra_context={"form": form, "pk": object.pk},
        )
        return f"""
        <div class="panel panel-default">
            <div class="panel-heading"><strong>Manage Interfaces with Excel</strong></div>
            <table class="table table-hover panel-body">
                <tbody>
                    <tr><td>
                        <a href='/plugins/uoft/interfaces-excel/{object.pk}/'>
                            <button class='btn btn-success'>Download '{object.name}-interfaces.xlsx'</button>
                        </a>
                    </td></tr>
                    <tr>
                        <td>
                            {form_html}
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>
        """
